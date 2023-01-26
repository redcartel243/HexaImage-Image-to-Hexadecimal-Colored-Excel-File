import openpyxl
from openpyxl.styles import PatternFill
from PIL import Image
import numpy as np
import pandas as pd

numberofimgs = int(input("how many images do you want to process?Number only"))
save_file_name = input("Give a name for result without color, without '.xlsx' please: ")
lastinput = input("Give a name for result with color")

names = {}
for i in range(0, numberofimgs):
    names[i] = input("Path to Image {}:".format(i + 1))
    print(names[i])
# check if the image name is valid or not. if valid, continue to the next step!
# use PIL for image manipulation
for i in range(0, numberofimgs):
    readmode = True
    while readmode:
        try:
            img = Image.open(names[i])
            readmode = False
        except:
            print("can not read ", names[i])
            print("please input Path to Image again\n ")
            names[i] = input("Path to Image {}:".format(i + 1))
    # resize the image
    basewidth = 100
    wpercent = (basewidth / float(img.size[0]))
    hsize = int((float(img.size[1]) * float(wpercent)))
    img = img.resize((basewidth, hsize), Image.ANTIALIAS)
    # convert the image into an array object
    # I also convert the image into a pandas DataFrame because I'm going to convert it into an excel file

    arr = np.array(img)
    arr_container = np.empty((arr.shape[0], arr.shape[1]), dtype=np.str)

    df = pd.DataFrame(arr_container)
    # convert the R,G,B into hexadecimal
    # populate the DataFrame with the hexadecimal value
    for l, x in enumerate(arr):
        for m, y in enumerate(x):
            print(len(arr[l][m]))
            if (len(arr[l][m]) == 3):
                r, g, b = arr[l][m]
                hexval = '%02x%02x%02x' % (r, g, b)
            else:
                r, g, b, c = arr[l][m]
                hexval = '%02x%02x%02x%02x' % (r, g, b, c)

            df[m][l] = hexval
    # save it! yeay!
    df.to_excel(str(save_file_name) + str(i) + ".xlsx")
    print("success!! now colour the file")

    # load excel with its path
    workbook = openpyxl.load_workbook(save_file_name + str(i) + ".xlsx")

    sheet = workbook.active

    for row_cells in sheet.iter_rows():
        for cell in row_cells:

            if (cell.value is None):
                continue
            st = str(cell.value)
            if (len(st) < 6):
                continue
            print(cell.value)
            """my_color = openpyxl.styles.colors.Color(rgb=str(cell.value))
            my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)
            cell.fill = my_fill
 
            sheet[cell.coordinate].fill = my_fill"""
            redFill = PatternFill(start_color=cell.value,
                                  end_color=cell.value,
                                  fill_type='solid')
            sheet[cell.coordinate].fill = redFill

    workbook.save(lastinput + str(i) + ".xlsx")
